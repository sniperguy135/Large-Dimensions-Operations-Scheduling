import os, re, sys
import csv
import openpyxl
import networkx as nx
import numpy as np

from string import ascii_uppercase as ABCD

import openpyxl.worksheet

""" DOCUMENTATION
data_input.py -- Data input handling and interpreting module
Data input subsystem

Methods to call:
MRP_input_validate
data_input_main

this documentation is kinda shoddy cos ive been through at least 3 or 4 iterations
yes i should probably have integrated better with the team and asked what the code was supposed to do
BEFORE starting to code
--------------

What the code does:
MRP_input_validate
1. Convert any .csv/.tsv into excel, stores in <same directory as> this file.
This is to take advantage of openpyxl's ability to read by column

2. Check data for MRP
At least 2 columns, 2 rows

3. Check if columns correspond to period, within 10 consecutive columns


main
1. Convert any .csv/.tsv into excel, stores in <same directory as> this file.

2. Check data 
Capacity (machines):
- Columns
    - Machine/Workcentre (str)
    - Count (int)
- No duplicates on machine column


3. Data validation - Check if the bare minimum for routing is present:
- Columns
    - Part
    - Operation
    - Component
    - Processing Time
    - Workcentre
- At least 1 row

4. Read the (generated) excel files and create a MultiDiGraph object based on the routing data
4.1 Check if BOM attributes within spec: at most 9 layers, at most 7 children for any node
4.2 Get unique list of (routing) machines from the excel 'Workcentre' column

5. Check data for MRP and capacity
MRP:
- Left most column: root node + all nodes with children
- At least 1 demand for end item (root node)

Capacity (machines):
- At least 1 of all machines referenced

6. Build Outputs

---------------------------

MAIN

Inputs:
file directories: str
    one for each, in order - routing, MRP, capacity
    e.g. "users/test/Desktop/routing.csv", "C:\\users\\test\\Desktop\\MRP.tsv", "/users/test/Desktop/capacity.xlsx", "C:\\users\\test\\Desktop\\im_bored.tsv"

    file types: .xlsx, .csv, .tsv
    NO TRAILING "/" or "\\" , since it denotes a directory instead of a file
    Extra backslash for windows directories since backslash is an escape character in Python

MRP start column: str
MRP end column: str    



Outputs:
Positive output 
- jobs dictionary
    keys = "process name - number" (str), value = dict 
        keys = 
            "Process time", value = (float),  
            "Machine", value = (str), 
- capacity dictionary
    key = workcentre (str), value = number of machines (int)
All in 1 tuple in order

-----------------------------------

Negative outputs:
- FileNotFoundError for file(s) not found
- TypeError for wrong/unhandled file type  
- ValueError for wrong data input
- ValueError for BOM attributes beyond spec
(Numbers to help with error identification on upper layers)
"""

#
# Main functions to call
#

def data_input_main(routing_directory:str, mrp_directory:str, capacity_directory:str, MRP_start_col: str, MRP_end_col: str):
    #
    # File type handling and verification
    #
    routing_directory = data_input_filetype_check(routing_directory, "routing")
    mrp_directory = data_input_filetype_check(mrp_directory, "MRP")
    capacity_directory = data_input_filetype_check(capacity_directory, "capacity")
    
    #
    # capacity contents verification
    #
    try:
        capacity_workbook = openpyxl.load_workbook(filename=capacity_directory, read_only=False, data_only=True) # readonlyworksheet has no iter_cols fn which i need
        capacity_worksheet = capacity_workbook[capacity_workbook.sheetnames[0]] # takes the first sheet
        # quick and dirty verifications
        assert capacity_worksheet.max_column == 2, "Should have 2 columns"
        assert set([x.lower() for x in [cell.value for cell in capacity_worksheet[1] if cell.value!=None]]) == set(["machine","count"]) or set([x.lower() for x in [cell.value for cell in capacity_worksheet[1] if cell.value!=None]]) == set(["workcentre","count"]), "Wrong header names"
        assert capacity_worksheet.max_row >= 2, "Should have at least 1 row for header and 1 row for 1 machine/workcentre"
        machine_or_workcentre_col_name = [x for x in [cell.value.lower() for cell in capacity_worksheet[1][0:2]] if x in ["machine", "workcentre"]]
        machine_or_workcentre_col_name = machine_or_workcentre_col_name[0] 
        machine_or_workcentre_col_idx = [cell.value.lower() for cell in capacity_worksheet[1]].index(machine_or_workcentre_col_name)
        capacity_machine_list = [x for x in [cell.value for cell in capacity_worksheet[["A","B"][machine_or_workcentre_col_idx]]] if x.lower() != machine_or_workcentre_col_name.lower() and x!=None]
        assert list(np.unique(capacity_machine_list)) == capacity_machine_list, "Duplicate rows for machines"
        count_col_idx = [cell.value.lower() for cell in capacity_worksheet[1]].index('count')
        count_col_values_ls = [x for x in [cell.value for cell in capacity_worksheet[["A","B"][count_col_idx]]] if x.lower() != 'count' and x!=None]
        assert 0 not in [int(x) for x in count_col_values_ls], "Machine count less than 1" # Check if counts at least 1
        assert len(count_col_values_ls) == len([x for x in count_col_values_ls if "." not in x]), "Machine count contains float" # Check if any full stops implying irrational number of machines
    except AssertionError as errAss:
        raise ValueError("Invalid capacity file contents", str(errAss))
    except AttributeError as errAttr:
        raise AttributeError("Capacity validation error", str(errAttr))

    #
    # routing contents verification
    #
    try:
        main_workbook = openpyxl.load_workbook(filename=routing_directory, read_only=True, data_only=True) # verifying only, so no write needed. faster and allows bigger files
        # print(main_workbook.sheetnames)
        main_worksheet = main_workbook[main_workbook.sheetnames[0]] # takes the first sheet
        # quick and dirty verifications
        # assert main_worksheet.max_column == 5, "Should only have 5 columns" # problematic check, sometimes deleted cells still count to column count
        assert [x for x in [cell.value for cell in main_worksheet[1] if cell.value!=None] if x.lower() not in ['part', 'operation', 'components required', 'processing time', 'workcentre']] == [], "Wrong header names"
        assert main_worksheet.max_row <= (7**14 + 1), "Total size overload" # worst case scenario, 7 children on all nodes on all layers
        assert main_worksheet.max_row >= 2, "No data" # minimum 1 parent with 1 child
    except AssertionError as errAss:
        raise ValueError("Invalid routing file contents", str(errAss))
    except AttributeError as errAttr:
        raise AttributeError("Routing validation error", str(errAttr))

    #
    # Graph building and verification
    #
    data_input_mdigraph, routing_machine_list = data_input_graph_builder(main_worksheet)
    root_node_name = data_input_graph_checker(data_input_mdigraph)

    # 
    # Verification of MRP with routing info
    #
    try:
        mrp_workbook = openpyxl.load_workbook(filename=mrp_directory, read_only=False, data_only=True) 
        mrp_worksheet = mrp_workbook[mrp_workbook.sheetnames[0]] 
        assert sorted([cell.value for cell in mrp_worksheet[root_node_name] if cell.value != None][1:]) == sorted(list(nx.topological_sort(data_input_mdigraph)))
    except AssertionError as errAss:
        raise ValueError("MRP file does not match routing contents", str(errAss))
    
    #
    # Verification of capacity with routing info
    #
    if sorted([x for x in routing_machine_list if x in capacity_machine_list]) != sorted(routing_machine_list):
        raise ValueError("Missing machines in capacity file")

    #
    # Find the jobs for the period
    #
    mrp_periods_column_letters = [MRP_start_col]
    next_letter = get_next_seq(MRP_start_col)
    while MRP_end_col not in mrp_periods_column_letters:
        mrp_periods_column_letters.append(next_letter)
        next_letter = get_next_seq(next_letter)

    # Since we only need the parts with operations, we can ignore rows for parts that do not have operations
    # Create a dictionary to map the part to the row number
    mrp_parts_row_dict = {}
    for mrp_row_num in range(2, mrp_worksheet.max_row+1):
        if mrp_worksheet[f"A{mrp_row_num}"].value == None: # beyond data range
            break        
        mrp_parts_row_dict[mrp_worksheet[f"A{mrp_row_num}"].value] = mrp_row_num

    #
    # Build outputs
    #
    jobs_dictionary = {} # BUILD

    # Find all the parent nodes
    # Then for each parent node, find out how many of the parent item needs to be made
    # Create that many number of each of the operations
    for parent_node in [node for node in data_input_mdigraph.nodes if data_input_mdigraph.out_degree(node) != 0]:
        number_of_parts_to_make = sum([int(mrp_worksheet[cell_ref].value) for cell_ref in [f'{column}{mrp_parts_row_dict[parent_node]}' for column in mrp_periods_column_letters] if mrp_worksheet[cell_ref].value != None])
        for child_node in layer_search(data_input_mdigraph, parent_node, only_layer=True):
            for edge_attr_dict in dict(data_input_mdigraph.get_edge_data(parent_node, child_node)).values():
                for rep in range(1, number_of_parts_to_make+1):
                    jobs_dictionary[f"{edge_attr_dict['operation']}-{rep}"] = {"Process time": edge_attr_dict['time'], "Machine": edge_attr_dict['machine']}

    # OLD METHOD: Since each edge represents a process, work from there
    """ for edge_info in list(data_input_mdigraph.edges):
        edge_attr_dict = data_input_mdigraph.edges[edge_info[0], edge_info[1], edge_info[2]]
        parent_node = edge_info[0]
        # Find the number of times to do the operation
        # This is stored in the number of parent items to make in the period range


        jobs_dictionary[edge_attr_dict['operation']] = {"Process time": edge_attr_dict['time'], "Machine": edge_attr_dict['machine']}"""
    
    # if job dictionary is empty then dont need to care, main.py should handle this
    
    capacity_dictionary = {} # BUILD
    for row in capacity_worksheet.iter_rows(min_row=2):
        machine_name, machine_count = [cell.value for cell in row]
        capacity_dictionary[machine_name] = int(machine_count)

    return jobs_dictionary, capacity_dictionary


def MRP_input_validate(mrp_validate_directory: str, start_col: str, end_col: str):
    # Check file type
    mrp_validate_directory = data_input_filetype_check(mrp_validate_directory, "MRP")
    
    # Open with openpyxl, basic validation
    try:
        mrp_map_workbook = openpyxl.load_workbook(mrp_validate_directory, data_only=True)
        mrp_map_worksheet = mrp_map_workbook[mrp_map_workbook.sheetnames[0]]
        # quick and dirty verifications
        assert mrp_map_worksheet.max_column >= 2, "Should have at least 2 columns"
        assert mrp_map_worksheet.max_row >= 2, "Should have at least 1 row for period and 1 row for 1 end item"
        int(mrp_map_worksheet["B1"].value)
    except AssertionError as errAss:
        raise ValueError("Invalid MRP file contents", str(errAss))
    
    # Check for period columns: if at most 10 consecutive columns, and if chosen columns have at least 1 job
    # if only 1 column, only need to check if the column is not empty
    start_col, end_col = start_col.upper(), end_col.upper()

    
    if [start_col,end_col] != sorted([start_col, end_col], key=lambda x: (len(x),x)):
        raise ValueError("Start column is after End column")
    elif len(start_col) > 3 or len(end_col) > 3:
        raise ValueError("Start or End columns invalid")
    elif start_col == "A" or end_col == "A":
        raise ValueError("Start/end column cannot be A, Column A is reserved for part letters. Refer to template MRP.")
    elif start_col == end_col and [cell.value for cell in mrp_map_worksheet[start_col] if cell.value != None][1:] != []:
        return mrp_validate_directory

    col_ls = [start_col]
    jobs_ls = []

    while col_ls[-1] != end_col:
        if mrp_map_worksheet[col_ls[-1]+"1"].value == None:
            raise ValueError("Columns selected do not have specified period. Refer to template MRP.")
        jobs_ls.append([cell.value for cell in mrp_map_worksheet[col_ls[-1]] if cell.value != None][1:])
        col_ls.append(get_next_seq(col_ls[-1]))

    if jobs_ls == []:
        raise ValueError("No jobs scheduled in specified columns")
    elif len(col_ls) > 10:
        raise ValueError("More than 10 columns chosen")
    else:
        return mrp_validate_directory


#
# Helper functions
#

def data_input_filetype_check(directory:str, data_name_filetype_check:str = ""):
    if directory[-4:] == ".csv" or directory[-4:] == ".tsv":
        directory = data_input_xlsx_convert(directory, data_name_filetype_check)
    elif directory[-5:] != ".xlsx":
        raise TypeError(f"Unsupported File Type - check if the {data_name_filetype_check} file is .xlsx, .csv or .tsv")
    return directory


def data_input_xlsx_convert(csv_tsv_directory:str, file_prefix_name:str = "test"): # Convert csv/tsv files to xlsx
    file_dir_splitter = "\\" if sys.platform == "win32" else "/"
    wbname = csv_tsv_directory.split(file_dir_splitter)[-1][:-4]

    file = open(csv_tsv_directory, 'r', encoding='utf-8-sig')
    if csv_tsv_directory[-4:] == ".csv":
        reader = csv.reader(file)  
    elif csv_tsv_directory[-4:] == ".tsv":
        reader = csv.reader(file, delimiter='\t') 
    else:
        raise TypeError(f"Unsupported File Type to convert - check if the {file_prefix_name} file is .csv or .tsv")
    
    # Make the 'Working files' directory if it does not exist
    try: 
        os.mkdir('Working files')
    except FileExistsError:
        pass

    dir_string = f"{file_dir_splitter}Working files{file_dir_splitter}{file_prefix_name}_{wbname}_converted.xlsx"
    if os.path.exists(dir_string): # if there is an existing file, delete it
        os.remove(dir_string)
    wb = openpyxl.Workbook() # create a new workbook to transfer contents to
    ws = wb.active
    for row in reader:
        ws.append(row)
    wb.save(os.getcwd()+dir_string)

    new_xlsx_directory = os.getcwd() + dir_string
    # print(new_xlsx_directory) # DEBUG STEP
    return new_xlsx_directory


def data_input_graph_builder(graph_build_worksheet: openpyxl.worksheet): # Converting into nx.DiGraph
    data_input_graph = nx.MultiDiGraph()
    try:
        header_list = [cell.value.lower() for cell in graph_build_worksheet[1] if cell.value != None]
    except AttributeError as errAttr:
        raise AttributeError("Graph builder error", str(errAttr))
    part_col = header_list.index('part')
    oper_col = header_list.index('operation')
    comp_col = header_list.index('components required')
    prtm_col = header_list.index('processing time')
    wcnt_col = header_list.index('workcentre')

    # Each node is a component
    # Child means that it relies on that part
    # Weight of arc shows how many of each child part is required: int
    # The edge "operation" attribute stores the operation which the child is required in: str
    # The edge "time" attribute stores the amount of time required for the operation: float
    # "time" should be the same for all edges of the same operation
    # The edge "machine" attribute stores the workcentre for the operation: str
    # types are not checked/enforced

    # Addon role: Give list of machine names (str)
    machine_list = []

    for row in graph_build_worksheet.iter_rows(min_row = 2): # read from second (first non-header) row onwards
        if row[part_col].value not in list(data_input_graph.nodes): # if part has not been added to graph
            data_input_graph.add_node(row[part_col].value)
        
        operation_name = str(row[oper_col].value)
        proc_time = float(row[prtm_col].value)
        machine_name = str(row[wcnt_col].value)
        if machine_name not in machine_list:
            machine_list.append(machine_name)

        for num_and_part in row[comp_col].value.split(" "): # to handle if there are multiple parts required for the operation
            # split into numerical and alphabetical components (3AB -> ["3", "AB"])
            weight_number, part_name = [x for x in re.split('(\d+)',num_and_part) if x != '']
            data_input_graph.add_edge(row[part_col].value, part_name, weight = weight_number, operation = operation_name, time = proc_time, machine = machine_name) 
            # old ways of adding edge attributes, absorbed into line above
            # data_input_graph[row[part_col].value][part_name]["operation"] = operation_name
            # data_input_graph[row[part_col].value][part_name]["time"] = proc_time
            # data_input_graph[row[part_col].value][part_name]["machine"] = machine_name
    
    return data_input_graph, machine_list


def data_input_graph_checker(graph_check: nx.MultiDiGraph): # Checking the BOM properties, returns root node (str)
    try:
        # https://stackoverflow.com/questions/4122390/getting-the-root-head-of-a-digraph-in-networkx-python
        assert len([n for n,d in graph_check.in_degree() if d==0]) == 1, "At least one unlinked child" #Check if only 1 root node (1 end product)
        root_node = list(nx.topological_sort(graph_check))[0]
        #print([node_name for node_name in graph_check.nodes if graph_check.succ[node_name] != {}])
        assert max(list(dict(enumerate(nx.bfs_layers(graph_check, root_node))).keys())) <= 14, "More than 14 layers" # Check max layer less than or equal to 14
        # comment: test above needs to pass to avoid expensive calculations below
        # print(layer_search(graph_check, root_node)) # for debugging to see what the graph looks like
        assert max([len(sub_dict) for sub_dict in list(layer_search(graph_check, root_node).values())]) <= 7, "More than 7 children" # Check max number of children for every node <= 7
    except AssertionError as errAss:
        raise ValueError("BOM attributes beyond spec", str(errAss))    
    return root_node


def layer_search(graph: nx.MultiDiGraph, nodename, networkdict: dict = {}, only_layer:bool = False): # finds all the nodes with children regardless of layer and stores it in a dictionary
    child_dict = dict(graph[nodename])
    if child_dict != {}:
        networkdict[nodename] = child_dict
        if only_layer:
            return child_dict.keys()        
        for childnodename in child_dict.keys():
            layer_search(graph, childnodename, networkdict)
    else:
        return
    return networkdict


def get_next_seq(prev_seq): # Used to get the next available alphabet. After 'Z', the next letter(s) will be 'AA', and after 'ZZ', 'AAA' and so on.
    # https://stackoverflow.com/a/78120210
    # '@' is the character before 'A' and is used as a zero padding
    # '[' is the character after 'Z', we add one to the end as an initial 'carryover'
    seq = list('@'+prev_seq.strip()+'[')
    for i in range(1,len(seq)):
        if seq[-i] == '[':
            seq[-i] = 'A'
            seq[-(i+1)] = chr(ord(seq[-(i+1)])+1)
    # Remove leading zero ('@') if still present
    seq = ''.join(seq).replace('@','')
    # Also remove the carryover character at the end
    return seq[:-1]



# Test running, only runs if this file is directly run
if __name__ == "__main__":
    None # placeholder in case all commented
    file_dir_splitter = "\\" if sys.platform == "win32" else "/"
    cwd_prefix = os.getcwd() + file_dir_splitter

    # check if can create a folder if it does not exist
    # data_input_filetype_check(cwd_prefix+"bicycle-routing.csv")
    job_dict_eg, capacity_dict_eg = data_input_main(cwd_prefix+"bicycle-routing.csv", cwd_prefix+"bicycle-MRP.csv", cwd_prefix+"bicycle-capacity.csv", "R", "AB")
    with open('egjobdict.txt', 'w') as f:
        print(job_dict_eg, file=f)
    with open('egcapdict.txt', 'w') as f:
        print(capacity_dict_eg, file=f)

